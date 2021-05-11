from telegram import Update, ReplyKeyboardMarkup, ReplyKeyboardRemove
from telegram.ext import Updater, MessageHandler, Filters, CommandHandler
from openpyxl import load_workbook



book = load_workbook('database.xlsx')
sheet_1 = book['Лист 1 - База данных']
time_page = book['Лист 1 - База данных']
TOKEN = "1623120001:AAFqRnhlwIAP_L4m7LGQvqiyE-jyU0cjZ9o"


days = [
    'понедельник',
    'вторник',
    'среда',
    'четверг',
    'пятница',
]


Понедельник = ''' 2  '''
Вторник = ''' 3 '''
Среда = '''4'''
Четверг = '''5'''
Пятница = ''' 6 '''


def main():
    updater = Updater(token=TOKEN)  # объект, который ловит сообщения из телеграм
    dispatcher = updater.dispatcher
    handler = MessageHandler(Filters.all, do_echo)  # отфильтровываем сообщения: теперь устройство должно реагировать
    start_handler = CommandHandler('start', do_start)
    help_handler = CommandHandler('help', do_help)
    keyboard_handler = MessageHandler(Filters.text, do_something)
    dispatcher.add_handler(start_handler)
    dispatcher.add_handler(help_handler)
    dispatcher.add_handler(keyboard_handler)
    dispatcher.add_handler(handler)
    updater.start_polling()
    updater.idle()


def do_echo(update: Update, context):
    user = update.message.from_user.is_bot
    name = update.message.from_user.first_name
    if user:
        update.message.reply_text(text=f"Ты - бот! Уходи отсюда!!!")
    else:
        update.message.reply_text(text=f"ААААА! {name} что ты делаешь?")
        update.message.reply_text(text="Я не понимаю")


def do_start(update, context):
    keyboard = [
        ["Понедельник", "Вторник", "Среда"],
        ["Четверг", "Пятница"]
    ]
    update.message.reply_text(text="Hallo!",
                              reply_markup=ReplyKeyboardMarkup(keyboard, one_time_keyboard=True,
                                                               resize_keyboard=True))  # текст сообщения, который бот автоматически будет выдавать
    update.message.reply_text(text="choose the day schedule for which you want",
                       reply_markup=ReplyKeyboardMarkup(keyboard, one_time_keyboard=True,
                                                        resize_keyboard=True))


def do_help(update, context):
    update.message.reply_text(text="что случилось? Всё ж было норм")


def do_something(update, context):
    text = update.message.text
    text = text.lower()

    if text in days:
        answer = get_schedule(text)
        update.message.reply_text(text=answer, reply_markup=ReplyKeyboardRemove())

    else:
        update.message.reply_text(text="Wrong day", reply_markup=ReplyKeyboardRemove())


def get_schedule(day):
    answer = ''
    for col in range(2, 7):
        if sheet_1.cell(row=3, column=col).value.lower() == day:
            break
    for i in range(3, sheet_1.max_row):
        cell = sheet_1.cell(row=i, column=col).value
        time = sheet_1.cell(row=i, column=7).value
        if cell is None:
            break
        answer += time + '  ' + cell + '\n'
    return answer

main()
